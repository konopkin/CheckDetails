import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox, QMainWindow 
import checkDetailInterface
from webbrowser import open as openWeb
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from requests import get as getPage
from openpyxl import load_workbook


ua = UserAgent()
header = {'User-Agent':str(ua.chrome)}

class InterfaceApp(QMainWindow, checkDetailInterface.Ui_MainWindow, QMessageBox):
	def __init__(self):
		global sheet
		global dataBase
		# Это здесь нужно для доступа к переменным, методам
		# и т.д. в файле design.py
		super().__init__()
		self.setupUi(self)
		#--------------------настройка кнопок--------------
		self.pushButton_openInBrowser.setEnabled(False)
		#-------------------------------------------------
		dataBase = load_workbook('./DataBase.xlsx')
		sheet = dataBase.active
		#------------------------------------привязка кнопок---------------
		self.pushButton_start.clicked.connect(self.StartClicked)
		self.pushButton_openInBrowser.clicked.connect(self.OpenUrlClicked)
		self.pushButton_SaveDB.clicked.connect(self.SaveToDB)
		self.pushButton_FindDB.clicked.connect(self.FindInBase)
		self.pushButton_SubtractDB.clicked.connect(self.SubtractDB)
		self.lineEdit_partNumber.textChanged.connect(self.PartTextChanged)
		#------------------------------------------------------------------

	def PartTextChanged(self):
		self.pushButton_openInBrowser.setEnabled(False)


	def StartClicked(self):
		global currentUrl
		currentUrl = self.HPParser(str(self.lineEdit_partNumber.text()))
		if ans == True:
			self.pushButton_openInBrowser.setEnabled(True)


	def OpenUrlClicked(self):
		openWeb(currentUrl)


	def SaveToDB(self):
		rowCount = sheet.max_row
		i = 2
		found = False
		while i <= rowCount:
			currentPartCheck = sheet["C" + str(i)].value
			if currentPartCheck == self.lineEdit_partNumber.text():
				sheet["A" + str(i)].value = int(sheet["A" + str(i)].value) + int(self.spin_count.text())
				if self.lineEdit_price.text() != None:
					sheet["B" + str(i)].value = self.lineEdit_price.text()
				found = True
				break
			i += 1
		if found == False:
			sheet.append([self.spin_count.text(), self.lineEdit_price.text(), str(self.lineEdit_partNumber.text())])
		dataBase.save('DataBase.xlsx')


	def FindInBase(self):
		rowCount = sheet.max_row
		i = 2
		found = False
		while i <= rowCount:
			currentPartCheck = sheet["C" + str(i)].value
			if currentPartCheck == self.lineEdit_partNumber.text():
				self.textEdit.setText(currentPartCheck)
				self.textEdit.append("Количество: "+ str(sheet["A" + str(i)].value))
				self.textEdit.append("Цена: "+ str(sheet["B" + str(i)].value))
				found = True
				break
			i += 1
		if found == False:
			self.textEdit.setText("Не найдено.")
			self.infoBox()
			
		dataBase.save('DataBase.xlsx')


	def SubtractDB(self):
	
		rowCount = sheet.max_row
		i = 2
		found = False
		while i <= rowCount:
			currentPartCheck = sheet["C" + str(i)].value
			if currentPartCheck == self.lineEdit_partNumber.text():
				sheet["A" + str(i)].value =int(sheet["A" + str(i)].value) - int(self.spin_count.text())
				found = True
				break
			i += 1
		if found == False:
			self.textEdit.setText("Этой позиции нет в базе")
			
		dataBase.save('DataBase.xlsx')

	def infoBox(self):
		global currentUrl
		infoBox = QMessageBox(self)
		infoBox.setIcon(QMessageBox.Warning)
		infoBox.setText("Не найдено")
		message = "В базе нет записей с данным Part номером. Искать по совместимым?"
		infoBox.setInformativeText(message)
		infoBox.setWindowTitle("Warning")
		infoBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
		reply = infoBox.exec_()
		if reply == 1024:
			rowCount = sheet.max_row
			try:
				currentUrl = self.HPParser(str(self.lineEdit_partNumber.text()))
				if ans == False:
					return
			except:
				return
			try:
				found = False
				currentPartMassive = connectedComps2.text.split("\n")
				for line in currentPartMassive:
					i = 2
					while i <= rowCount:
						currentPartCheck = sheet["C" + str(i)].value
						if currentPartCheck == line:
							self.textEdit.append("-------------")
							self.textEdit.append(currentPartCheck)
							self.textEdit.append("Количество: ")
							self.textEdit.append(str(sheet["A" + str(i)].value))
							self.textEdit.append("Цена: ")
							self.textEdit.append(str(sheet["B" + str(i)].value))
							found = True
							break
						i += 1
				if found == False:
					self.textEdit.append("---------\nПоиск не дал результатов.")

			except:
				found = False
				currentPartMassive = connectedPartNums.text.split("\n")
				for line in currentPartMassive:
					i = 2
					while i <= rowCount:
						currentPartCheck = sheet["C" + str(i)].value
						if currentPartCheck == line:
							self.textEdit.append("-------------")
							self.textEdit.append(currentPartCheck)
							self.textEdit.append("Количество: ")
							self.textEdit.append(str(sheet["A" + str(i)].value))
							self.textEdit.append("Цена: ")
							self.textEdit.append(str(sheet["B" + str(i)].value))


							found = True
							break
						i += 1
			if ans == True:
				self.pushButton_openInBrowser.setEnabled(True)


#-------------------HPParser----------
	def HPParser(self, partNumber):
		global ans
		global connectedComps2
		global connectedPartNums
		BASE_URL = "https://partsnb.ru"
		page = getPage("%s/search/%s" % (BASE_URL, partNumber))
		soup = BeautifulSoup(page.text, 'html.parser')
		try:
			firstSoup = soup.find("span", class_ = "categoryTitle").find("a", recursive=False).get('href')
			ans = True
		except AttributeError:
			self.textEdit.setText("Invalid Part Number")
			ans = False
			return
		page = getPage("%s%s" % (BASE_URL, firstSoup))
		soup = BeautifulSoup(page.text, 'html.parser')
		firstSoup = soup.find("a", class_ = "link-product").get('href')
		page = getPage("%s%s" % (BASE_URL, str(firstSoup)))
		soup = BeautifulSoup(page.text, 'html.parser')
		connectedComps = soup.find("div", class_ = "product-info").findAll("ul")[0]
		connectedPartNums = soup.find("div", class_ = "product-info").findAll("ul")[1]
		self.textEdit.setText(soup.find("div", class_ = "page-body").findAll("h1")[0].text)

		try: #если там три ul и кнопка показать все
			connectedComps2 = soup.find("div", class_ = "product-info").findAll("ul")[2]
			self.textEdit.append("Совместимые компьютеры:")
			self.textEdit.append(connectedComps.text)
			self.textEdit.append(connectedPartNums.text)
			self.textEdit.append("\n%s\nСовместимые Part Numbers:" % ("*"*45))
			self.textEdit.append(connectedComps2.text)

		except:
			connectedComps2 = None
			self.textEdit.append("Совместимые компьютеры:")
			self.textEdit.append(connectedComps.text)
			self.textEdit.append("\n%s\nСовместимые Part Numbers:" % ("*"*45))
			self.textEdit.append(connectedPartNums.text)
		
		return ("%s%s" % (BASE_URL, str(firstSoup)))
#----------------------------------------------------------------------------------------


def main():
    app = QApplication(sys.argv)  # Новый экземпляр QApplication
    window = InterfaceApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно
    app.exec_()

if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем

	main()